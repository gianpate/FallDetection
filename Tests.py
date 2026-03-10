import os
import csv
from utils.parsers.JSONparsers import basicParser
from openpyxl import Workbook
from openpyxl.styles import Font



def write_markdown_report(table, recognizers, filename):

    stats = []

    for r_index in range(len(recognizers)):
        TP = FP = TN = FN = 0

        for row in table:
            expected = row[1]
            detected = row[r_index + 2][0]  # +2 because:
                                            # 0 = sample
                                            # 1 = expected

            if detected and expected:
                TP += 1
            elif detected and not expected:
                FP += 1
            elif not detected and not expected:
                TN += 1
            else:
                FN += 1

        stats.append((TP, FP, TN, FN))


    #  Write file
    with open(filename, "w") as f:
        f.write("## Recognizer Summary\n\n")

        summary_header = ["Recognizer", "TP", "FP", "TN", "FN"]
        summary_separator = [":--", "--:", "--:", "--:", "--:"]

        f.write("| " + " | ".join(summary_header) + " |\n")
        f.write("| " + " | ".join(summary_separator) + " |\n")

        for recog, (TP, FP, TN, FN) in zip(recognizers, stats):
            f.write(f"| {recog.__name__} | {TP} | {FP} | {TN} | {FN} |\n")

        f.write("\n\n")


        # Build header rows
        header = ["SAMPLE"]
        for recog in recognizers:
            name = recog.__name__
            header.append(f"{name} SCORE")
            header.append(f"{name} PASS")


        #  Build separator row (alignment row)
        #    Left for SAMPLE and PASS
        #    Right for SCORE

        separator = [":--"]  # SAMPLE left aligned
        for _ in recognizers:
            separator.append("--:")   # SCORE right
            separator.append(":--")   # PASS left

        # First header row
        f.write("| " + " | ".join(header) + " |\n")
        # Alignment separator
        f.write("| " + " | ".join(separator) + " |\n")
 
        # Data rows
        for row in table:
            sample = row[0]
            expected = row[1]

            md_row = [sample]

            for (detected, score) in row[2:]:
                status = "🟢 PASS" if detected == expected else "🔴 FAIL"
                md_row.append(f"{score:.1f}")
                md_row.append(status)

            f.write("| " + " | ".join(md_row) + " |\n")



def write_xlsx_report(table, recognizers, filename):

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Headers (two rows)
    header_row1 = ["SAMPLE"]
    header_row2 = [""]
    for recog in recognizers:
        header_row1.append(recog.__name__)
        header_row1.append("")
        header_row2.append("PASS")
        header_row2.append("SCORE")
    ws.append(header_row1)
    ws.append(header_row2)

    # Merge recognizer header cells
    col = 2
    for _ in recognizers:
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col+1)
        col += 2

    #  Data
    green_font = Font(color="008000")  # green
    red_font = Font(color="FF0000")    # red

    for row in table:
        excel_row = [row[0]]
        for (p, s) in row[1:]:
            excel_row.append("PASS" if p else "FAIL")
            excel_row.append(s)
        ws.append(excel_row)

        # Color PASS/FAIL 
        last_row = ws.max_row
        col_index = 2

        for (p, _) in row[1:]:
            cell = ws.cell(row=last_row, column=col_index)
            if p:
                cell.font = green_font
            else:
                cell.font = red_font
            col_index += 2
    wb.save(filename)



def runRecognizersTest(recognizers, fallSamples, noFallSamples,
                       md_output="testResults/recognizersResults.md",
                       xlsx_output="testResults/recognizersResults.xlsx"):

    # Build unified sample list with ground truth
    samples = []
    for sample in fallSamples:
        samples.append((sample, True))   # True = Fall expected
    for sample in noFallSamples:
        samples.append((sample, False))  # False = NoFall expected


    # Collect structured results
    # [sample_name, (pass_bool, score), (pass_bool, score), ...]
    table = []
    for sample, expected in samples:
        joints = basicParser(sample)

        row = [sample, expected] 
    
        for recog in recognizers:
            detected, score = recog(joints)
            row.append((detected, score))
        table.append(row)


    # Write TXT aligned table
    write_markdown_report(table, recognizers, md_output)
    # Write CSV (Excel friendly)
    # write_xlsx_report(table, recognizers, xlsx_output)

    print(f"Reports written to: {md_output} and {xlsx_output}")




def write_markdown_detector_report(table, detectors, filename):
    """
    Writes a markdown report for boolean-only detectors.
    table: list of rows, each row = [sample, expected, det1_result, det2_result, ...]
    detectors: list of detector functions
    """
    stats = []
    for d_index in range(len(detectors)):
        TP = FP = TN = FN = 0
        for row in table:
            expected = row[1]
            detected = row[d_index + 2]  # +2 because index 0=sample, 1=expected
            if detected and expected:
                TP += 1
            elif detected and not expected:
                FP += 1
            elif not detected and not expected:
                TN += 1
            else:
                FN += 1
        stats.append((TP, FP, TN, FN))

    with open(filename, "w") as f:
        # Summary table
        f.write("## Detector Summary\n\n")
        header = ["Detector", "TP", "FP", "TN", "FN"]
        separator = [":--", "--:", "--:", "--:", "--:"]
        f.write("| " + " | ".join(header) + " |\n")
        f.write("| " + " | ".join(separator) + " |\n")
        for det, (TP, FP, TN, FN) in zip(detectors, stats):
            f.write(f"| {det.__name__} | {TP} | {FP} | {TN} | {FN} |\n")
        f.write("\n\n")

        # Detailed table
        detail_header = ["SAMPLE"] + [det.__name__ for det in detectors]
        detail_separator = [":--"] + [":--" for _ in detectors]
        f.write("| " + " | ".join(detail_header) + " |\n")
        f.write("| " + " | ".join(detail_separator) + " |\n")

        for row in table:
            sample = row[0]
            expected = row[1]
            md_row = [sample]
            for detected in row[2:]:
                status = "🟢 PASS" if detected == expected else "🔴 FAIL"
                md_row.append(status)
            f.write("| " + " | ".join(md_row) + " |\n")


def run_FallDetectorsTest(detectors, fallSamples, noFallSamples,
                          md_output="testResults/fallDetectorsResults.md",
                          xlsx_output="testResults/fallDetectorsResults.xlsx"):
    """
    Tests a list of boolean‑only fall detectors on positive (fall) and negative (no fall) samples.
    Generates markdown and Excel reports with PASS/FAIL and confusion matrix statistics.
    """
    # Build unified sample list with ground truth
    samples = []
    for sample in fallSamples:
        samples.append((sample, True))   # fall expected
    for sample in noFallSamples:
        samples.append((sample, False))  # no fall expected

    # Collect structured results
    # table entry: [sample_name, expected, det1_result, det2_result, ...]
    table = []
    for sample, expected in samples:
        joints = basicParser(sample)          # assumes this function returns joint data
        row = [sample, expected]
        for det in detectors:
            detected = det(joints)             # detector returns bool only
            row.append(detected)
        table.append(row)

    # Write reports
    write_markdown_detector_report(table, detectors, md_output)
    # write_xlsx_detector_report(table, detectors, xlsx_output)

    print(f"Reports written to: {md_output} and {xlsx_output}")